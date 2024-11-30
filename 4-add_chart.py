from openpyxl import load_workbook
from openpyxl.chart import BarChart, Reference
# 1 - Lê pasta de trabalho e planilha
wb = load_workbook('Importando-dados-com-python\pivot_table.xlsx')
sheet = wb['Relatório']

# 2 - Referências de linhas e colnas
min_column = wb.active.min_column
max_column = wb.active.min_column
min_row = wb.active.min_row
max_row = wb.active.max_row

# 3 - Adcionando Dados e Categorias no Gráfico
bar_chart = BarChart()

categories = Reference(
    sheet,
    min_col=min_column + 1,
    max_col=min_column,
    min_row=min_row,
    max_row=max_row
)

bar_chart.add_data(data, titles_from_data=True)
bar_chart.set_categories(categories)
# 4 - Criango grágico
sheet.add_chart(bar_chart, 'B10')
bar_chart.title = 'Vendas por Fabricantes'
bar_chart.style = 2


# 5 - Salvando o workbook
wb.save('Importando-dados-com-python\barchart.xlsx')